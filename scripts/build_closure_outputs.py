import csv
import json
import re
import shutil
import tempfile
import zipfile
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


# Build a closure dataset from the best files currently on hand:
# 1. The cumulative federal closure workbook as of 2025-04-03.
# 2. The September 2024 helper workbook with both PEPS closures and IPEDS exits.
# 3. Monthly closed-school report zips for newer federal updates.
#
# The output is a running event list plus three derived tabs:
# - main campus closures
# - branch campus closures
# - mergers / consolidations
#
# We rely on PEPS/FSA-derived sources for closure events and use IPEDS to enrich
# those events with sector, category, and latest student headcount.

REPO_ROOT = Path(__file__).resolve().parents[1]

BASELINE_CLOSURE_XLSX = REPO_ROOT / "closure_file_20250403.xlsx"
SEPT_2024_CLOSURE_XLSX = Path(r"C:\Users\mv3031\Downloads\closures_sep24.xlsx")
MONTHLY_SECTION_1_ZIP = Path(r"C:\Users\mv3031\Downloads\MONTHLY_CLOSED_REPORT_SECTION_1_2026040107063586645.zip")
MONTHLY_SECTION_2_ZIP = Path(r"C:\Users\mv3031\Downloads\MONTHLY_CLOSED_REPORT_SECTION_2_2026040107103586655.zip")
ACCREDITATION_ACTIONS_CSV = REPO_ROOT / "accreditation" / "accreditation_tracker_actions_joined.csv"
IPEDS_CANONICAL_CSV = REPO_ROOT / "ipeds" / "ipeds_financial_health_dataset_2014_2024.csv"
IPEDS_RAW_CSV = REPO_ROOT / "ipeds" / "ipeds_financial_health_raw_2014_2024.csv"
IPEDS_DOWNLOADS_DIR = Path(r"C:\Users\mv3031\Downloads")

OUTPUT_DIR = REPO_ROOT / "federal_closure"
RUNNING_OUTPUT = OUTPUT_DIR / "running_closures.csv"
MAIN_OUTPUT = OUTPUT_DIR / "main_campus_closures.csv"
BRANCH_OUTPUT = OUTPUT_DIR / "branch_campus_closures.csv"
MERGER_OUTPUT = OUTPUT_DIR / "mergers_consolidations.csv"
PRIVATE_FEDERAL_MAIN_OUTPUT = OUTPUT_DIR / "private_sector_federal_main_closures.csv"
SUMMARY_OUTPUT = OUTPUT_DIR / "closure_pipeline_summary.json"
MIN_EVENT_YEAR = 2008
MAX_EVENT_YEAR = date.today().year


def normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def safe_int(value):
    if value in (None, "", "NA"):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def sector_from_sch_type(sch_type):
    return {
        1: "Public",
        2: "Private not-for-profit",
        3: "Private for-profit",
    }.get(sch_type, "")


def slug_key(*parts):
    return "|".join(normalize_text(part).lower() for part in parts if part not in (None, ""))


def read_csv_rows(path):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_zipped_csv_rows(path):
    with zipfile.ZipFile(path) as zf:
        name = [member for member in zf.namelist() if member.lower().endswith((".csv", ".txt"))][0]
        with zf.open(name, "r") as handle:
            text = handle.read().decode("utf-8-sig", errors="ignore").splitlines()
    return list(csv.DictReader(text))


def load_annual_ipeds_file_maps():
    # For closure enrichment we want the last pre-closure IPEDS snapshot rather
    # than only the newest year. That lets us fill sector, category, and
    # headcount for institutions that disappeared from later directories.
    annual_hd_by_year = {}
    annual_effy_by_year = {}
    annual_opeid8_to_unitid = {}
    annual_opeid6_to_unitid = {}
    annual_hd_rows_by_unitid = defaultdict(list)

    for year in range(2008, 2025):
        hd_zip = IPEDS_DOWNLOADS_DIR / f"HD{year}.zip"
        if hd_zip.exists():
            rows = read_zipped_csv_rows(hd_zip)
            annual_hd_by_year[year] = {}
            for row in rows:
                unitid = normalize_text(row.get("UNITID"))
                if not unitid:
                    continue
                annual_hd_by_year[year][unitid] = row
                annual_hd_rows_by_unitid[unitid].append((year, row))
                opeid8 = normalize_text(row.get("OPEID"))
                if opeid8:
                    annual_opeid8_to_unitid.setdefault(opeid8, unitid)
                    annual_opeid6_to_unitid.setdefault(opeid8[:6].lstrip("0"), unitid)

        effy_zip = IPEDS_DOWNLOADS_DIR / f"EFFY{year}.zip"
        if effy_zip.exists():
            rows = read_zipped_csv_rows(effy_zip)
            annual_effy_by_year[year] = {}
            for row in rows:
                unitid = normalize_text(row.get("UNITID"))
                if not unitid:
                    continue
                effy_alev = normalize_text(row.get("EFFYALEV"))
                lstudy = normalize_text(row.get("LSTUDY"))
                # Older EFFY files often do not carry EFFYALEV, but they do
                # provide duplicated total rows where LSTUDY 999 is the
                # institution-level total. Newer files use EFFYALEV 1.
                is_total_row = (
                    effy_alev == "1"
                    or (not effy_alev and lstudy in {"999", "1"})
                )
                if not is_total_row:
                    continue
                headcount = safe_int(row.get("EFYTOTLT"))
                if headcount is not None and unitid not in annual_effy_by_year[year]:
                    annual_effy_by_year[year][unitid] = headcount

    return annual_hd_by_year, annual_effy_by_year, annual_opeid8_to_unitid, annual_opeid6_to_unitid, annual_hd_rows_by_unitid


def fallback_sector_label(control_value):
    return {
        "1": "Public",
        "2": "Private not-for-profit",
        "3": "Private for-profit",
    }.get(str(control_value).strip(), "")


def fallback_category_label(iclevel_value):
    return {
        "1": "4-year or above",
        "2": "At least 2 but less than 4 years",
        "3": "Less than 2 years",
    }.get(str(iclevel_value).strip(), "")


def load_ipeds_lookups():
    canonical_rows = read_csv_rows(IPEDS_CANONICAL_CSV)
    raw_rows = read_csv_rows(IPEDS_RAW_CSV)
    annual_hd_by_year, annual_effy_by_year, annual_opeid8_to_unitid, annual_opeid6_to_unitid, annual_hd_rows_by_unitid = load_annual_ipeds_file_maps()

    rows_by_unitid = defaultdict(list)
    opeid8_to_unitid = {}
    opeid6_to_unitid = {}
    for row in canonical_rows:
        unitid = normalize_text(row.get("unitid"))
        if unitid:
            rows_by_unitid[unitid].append(row)

    for unitid_rows in rows_by_unitid.values():
        unitid_rows.sort(key=lambda r: safe_int(r.get("year")) or -1)

    for row in raw_rows:
        unitid = normalize_text(row.get("unitid"))
        opeid8 = normalize_text(row.get("opeid"))
        if not unitid or not opeid8:
            continue
        opeid8_to_unitid.setdefault(opeid8, unitid)
        opeid6 = opeid8[:6].lstrip("0")
        if opeid6:
            opeid6_to_unitid.setdefault(opeid6, unitid)

    for opeid8, unitid in annual_opeid8_to_unitid.items():
        opeid8_to_unitid.setdefault(opeid8, unitid)
    for opeid6, unitid in annual_opeid6_to_unitid.items():
        if opeid6:
            opeid6_to_unitid.setdefault(opeid6, unitid)

    return rows_by_unitid, opeid8_to_unitid, opeid6_to_unitid, annual_hd_by_year, annual_effy_by_year, annual_hd_rows_by_unitid


def fallback_year_candidates(event_year):
    # Use the last pre-event snapshot first. If the institution is already gone
    # from that year, walk backward through older IPEDS files.
    target = safe_int(event_year)
    if target is None:
        target = MAX_EVENT_YEAR
    start_year = min(max(target - 1, 2008), 2024)
    return list(range(start_year, 2007, -1))


def build_ipeds_status_index(annual_hd_rows_by_unitid):
    # Track first/last directory appearances so we can tell when a school drops
    # off the annual IPEDS universe and what status code it carried last.
    status_index = {}
    for unitid, rows in annual_hd_rows_by_unitid.items():
        rows = sorted(rows, key=lambda item: item[0])
        first_year, _ = rows[0]
        last_year, last_row = rows[-1]
        status_index[unitid] = {
            "first_year": first_year,
            "last_year": last_year,
            "last_act": normalize_text(last_row.get("ACT")),
            "last_closedat": normalize_text(last_row.get("CLOSEDAT")),
            "last_name": normalize_text(last_row.get("INSTNM")),
            "last_state": normalize_text(last_row.get("STABBR")),
            "last_opeid8": normalize_text(last_row.get("OPEID")),
        }
    return status_index


def parse_closedat_year(value):
    value = normalize_text(value)
    if not value or value in {"-2", "-1"}:
        return None
    match = re.search(r"(\d{4})$", value)
    return safe_int(match.group(1)) if match else None


def find_annual_hd_row(unitid, event_year, annual_hd_by_year):
    for year in fallback_year_candidates(event_year):
        row = annual_hd_by_year.get(year, {}).get(unitid)
        if row:
            return year, row
    return None, None


def find_annual_headcount(unitid, event_year, annual_effy_by_year):
    for year in fallback_year_candidates(event_year):
        headcount = annual_effy_by_year.get(year, {}).get(unitid)
        if headcount is not None:
            return year, headcount
    return None, None


def pick_ipeds_row(unitid, event_year, rows_by_unitid):
    rows = rows_by_unitid.get(unitid, [])
    if not rows:
        return None

    target_year = safe_int(event_year)
    eligible = [
        row for row in rows
        if safe_int(row.get("year")) is not None and (target_year is None or safe_int(row.get("year")) <= target_year)
    ]

    # Use the latest pre-event observation when possible. If none exist,
    # fall back to the latest available row in the canonical dataset.
    chosen = eligible[-1] if eligible else rows[-1]

    # If the chosen row has no headcount but an earlier one does, use the last
    # pre-event row with a nonmissing headcount so the workbook stays useful.
    if chosen and safe_int(chosen.get("enrollment_headcount_total")) is None:
        for row in reversed(eligible or rows):
            if safe_int(row.get("enrollment_headcount_total")) is not None:
                chosen = row
                break

    return chosen


def enrich_record(record, rows_by_unitid, opeid8_to_unitid, opeid6_to_unitid, annual_hd_by_year, annual_effy_by_year, ipeds_status_index):
    opeid8 = normalize_text(record.get("opeid8"))
    opeid6 = normalize_text(record.get("opeid6"))
    unitid = normalize_text(record.get("unitid"))

    if not unitid and opeid8:
        unitid = opeid8_to_unitid.get(opeid8, "")
        if unitid:
            record["unitid_match_method"] = "opeid8"
    if not unitid and opeid6:
        unitid = opeid6_to_unitid.get(opeid6, "")
        if unitid:
            record["unitid_match_method"] = "opeid6"

    record["unitid"] = unitid or ""

    ipeds_row = pick_ipeds_row(unitid, record.get("event_year"), rows_by_unitid) if unitid else None
    annual_hd_year, annual_hd_row = find_annual_hd_row(unitid, record.get("event_year"), annual_hd_by_year) if unitid else (None, None)
    annual_headcount_year, annual_headcount = find_annual_headcount(unitid, record.get("event_year"), annual_effy_by_year) if unitid else (None, None)

    if ipeds_row:
        record["institution_name_ipeds"] = normalize_text(ipeds_row.get("institution_name"))
        record["state"] = normalize_text(record.get("state") or ipeds_row.get("state"))
        record["sector"] = normalize_text(ipeds_row.get("control_label"))
        record["institutional_category"] = normalize_text(ipeds_row.get("category"))
        record["tracker_category"] = normalize_text(ipeds_row.get("category"))
        record["institution_status"] = normalize_text(ipeds_row.get("institution_status"))
        record["is_active"] = normalize_text(ipeds_row.get("is_active"))
        record["student_headcount"] = safe_int(ipeds_row.get("enrollment_headcount_total"))
        record["student_headcount_year"] = safe_int(ipeds_row.get("year"))

    if not record.get("sector") and safe_int(record.get("sch_type")) is not None:
        record["sector"] = sector_from_sch_type(safe_int(record.get("sch_type")))

    if annual_hd_row:
        if not record.get("institution_name_ipeds"):
            record["institution_name_ipeds"] = normalize_text(annual_hd_row.get("INSTNM"))
        record["state"] = normalize_text(record.get("state") or annual_hd_row.get("STABBR"))
        record["sector"] = record.get("sector") or fallback_sector_label(annual_hd_row.get("CONTROL"))
        if not record.get("institutional_category"):
            record["institutional_category"] = fallback_category_label(annual_hd_row.get("ICLEVEL"))
        if not record.get("tracker_category"):
            record["tracker_category"] = ""
        if not record.get("is_active"):
            record["is_active"] = normalize_text(annual_hd_row.get("ACT"))
        if not record.get("student_headcount_year") and annual_hd_year is not None and annual_headcount is None:
            record["student_headcount_year"] = annual_hd_year

    if annual_headcount is not None and not record.get("student_headcount"):
        record["student_headcount"] = annual_headcount
        record["student_headcount_year"] = annual_headcount_year

    if not record.get("tracker_category") and record.get("institutional_category"):
        record["tracker_category"] = record.get("institutional_category")

    unitid_status = ipeds_status_index.get(unitid) if unitid else None
    record["ipeds_first_seen_year"] = unitid_status.get("first_year") if unitid_status else None
    record["ipeds_last_seen_year"] = unitid_status.get("last_year") if unitid_status else None
    record["ipeds_last_act"] = unitid_status.get("last_act") if unitid_status else ""
    record["ipeds_last_closedat"] = unitid_status.get("last_closedat") if unitid_status else ""

    return record


def classify_opeid_record(opeid8, default_type=""):
    opeid8 = normalize_text(opeid8)
    if opeid8 and opeid8.endswith("00"):
        return "main_campus"
    if opeid8:
        return "branch_campus"
    return default_type


def read_baseline_closure_file():
    workbook = load_workbook(BASELINE_CLOSURE_XLSX, read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        school_name, state, opeid6, sch_type, close_year, closed_count, pgm_length = row
        if opeid6 is None or close_year is None:
            continue
        rows.append({
            "source_group": "baseline_closure_file",
            "source_file": BASELINE_CLOSURE_XLSX.name,
            "source_section": "",
            "event_label": "Federal closure workbook baseline",
            "school_name_source": normalize_text(school_name),
            "state": normalize_text(state),
            "opeid8": "",
            "opeid6": str(int(opeid6)),
            "unitid": "",
            "event_date": "",
            "event_year": safe_int(close_year),
            "record_kind": "closure",
            "classification": "main_campus",
            "classification_basis": "Cumulative federal closure workbook row (full-system closure logic).",
            "sch_type": safe_int(sch_type),
            "closed_count": safe_int(closed_count),
            "pgm_length": safe_int(pgm_length),
            "source_title": "",
            "source_url": "",
            "notes": "",
        })
    return rows


def read_sep24_workbook():
    workbook = load_workbook(SEPT_2024_CLOSURE_XLSX, read_only=True, data_only=True)

    peps_rows = []
    ws_peps = workbook["PEPS closures in 2024"]
    for row in ws_peps.iter_rows(min_row=2, values_only=True):
        close_date, opeid, school_name, location, address, city, state, zip_code, country = row
        if not opeid:
            continue
        opeid8 = normalize_text(opeid)
        close_date = normalize_text(close_date)
        year = safe_int(close_date[:4]) if close_date else None
        peps_rows.append({
            "source_group": "peps_closures_2024_sheet",
            "source_file": SEPT_2024_CLOSURE_XLSX.name,
            "source_section": "PEPS closures in 2024",
            "event_label": "PEPS closure row",
            "school_name_source": normalize_text(school_name),
            "state": normalize_text(state),
            "opeid8": opeid8,
            "opeid6": opeid8[:6].lstrip("0"),
            "unitid": "",
            "event_date": close_date,
            "event_year": year,
            "record_kind": "closure",
            "classification": classify_opeid_record(opeid8),
            "classification_basis": "PEPS closure row from September 2024 workbook.",
            "sch_type": None,
            "closed_count": None,
            "pgm_length": None,
            "source_title": "",
            "source_url": "",
            "notes": normalize_text(location),
        })

    closing_rows = []
    ws_closing = workbook["Colleges listed as closing"]
    for row in ws_closing.iter_rows(min_row=3, values_only=True):
        unitid, name, state = row
        if not unitid:
            continue
        closing_rows.append({
            "unitid": str(int(unitid)),
            "school_name_source": normalize_text(name),
            "state": normalize_text(state),
            "event_year": 2024,
            "source_file": SEPT_2024_CLOSURE_XLSX.name,
            "source_section": "Colleges listed as closing",
        })

    opening_rows = []
    ws_opening = workbook["Colleges listed as opening"]
    for row in ws_opening.iter_rows(min_row=3, values_only=True):
        unitid, name, state = row
        if not unitid:
            continue
        opening_rows.append({
            "unitid": str(int(unitid)),
            "school_name_source": normalize_text(name),
            "state": normalize_text(state),
        })

    return peps_rows, closing_rows, opening_rows


def parse_monthly_pdf_zip(zip_path, section_label):
    records = []
    with zipfile.ZipFile(zip_path) as zf:
        pdf_name = zf.namelist()[0]
        with tempfile.TemporaryDirectory() as temp_dir:
            pdf_path = Path(temp_dir) / pdf_name
            pdf_path.write_bytes(zf.read(pdf_name))
            reader = PdfReader(str(pdf_path))

            for page in reader.pages[1:]:
                text = page.extract_text() or ""
                lines = [normalize_text(line) for line in text.splitlines() if normalize_text(line)]
                if not lines:
                    continue
                # Each page is one school record. Some begin with a region line,
                # so prefer the second line in that case.
                school_name = lines[1] if lines[0].startswith("Region ") and len(lines) > 1 else lines[0]
                opeid_match = re.search(r"OPE ID:\s*(\d{8})", text)
                closure_match = re.search(r"Closure Date:\s*(\d{2}/\d{2}/\d{2})", text)
                if not opeid_match or not closure_match:
                    continue
                opeid8 = opeid_match.group(1)
                close_date = closure_match.group(1)
                year = 2000 + int(close_date[-2:])
                records.append({
                    "source_group": "monthly_closed_school_report",
                    "source_file": zip_path.name,
                    "source_section": section_label,
                    "event_label": "Monthly closed school report",
                    "school_name_source": school_name,
                    "state": "",
                    "opeid8": opeid8,
                    "opeid6": opeid8[:6].lstrip("0"),
                    "unitid": "",
                    "event_date": close_date,
                    "event_year": year,
                    "record_kind": "closure",
                    "classification": classify_opeid_record(opeid8),
                    "classification_basis": f"{section_label} monthly federal closed-school report.",
                    "sch_type": None,
                    "closed_count": None,
                    "pgm_length": None,
                    "source_title": "",
                    "source_url": "",
                    "notes": "",
                })
    return records


def classify_accreditation_action(action_text):
    text = normalize_text(action_text)
    lower = text.lower()

    if not text:
        return None
    if "receiving institution" in lower:
        return None
    if "open a branch campus" in lower or "affirmed the quality of the branch campus" in lower:
        return None

    # Exclude teach-outs that are clearly about academic programs rather than
    # an institution, branch campus, or additional location.
    program_markers = [
        "certificate",
        "associate of",
        "bachelor of",
        "master of",
        "doctor of",
        "degree program",
        "degrees",
        "programs",
        "program.",
        "program,",
        "curriculum",
    ]
    location_markers = ["main campus", "branch campus", "additional location", "location:", "locations:", "campus:"]
    has_location_marker = any(marker in lower for marker in location_markers)
    if any(marker in lower for marker in program_markers) and not has_location_marker and "closure of " not in lower:
        return None

    if "close the main campus" in lower or "closing of a main campus" in lower:
        return "main_campus"
    if "close a branch campus" in lower or "closing of a branch campus" in lower or "teach-out of a branch campus" in lower or "teach out of a branch campus" in lower:
        return "branch_campus"
    if "close an additional location" in lower or "close the additional location" in lower:
        return "branch_campus"
    if "teach-out of an additional location" in lower or "teach out of an additional location" in lower:
        return "branch_campus"
    if "teach-out of the additional location" in lower or "teach out of the additional location" in lower:
        return "branch_campus"
    if "teach out students at an additional location" in lower or "teach-out students at an additional location" in lower:
        return "branch_campus"
    if "teach out students at the branch campus" in lower or "teach-out students at the branch campus" in lower:
        return "branch_campus"
    if "teach out of students enrolled at the following additional location" in lower or "teach-out of students enrolled at the following additional location" in lower:
        return "branch_campus"
    if "teach out of students enrolled at the additional location" in lower or "teach-out of students enrolled at the additional location" in lower:
        return "branch_campus"
    if "teach-out plan for closing" in lower and ("additional location" in lower or "locations" in lower):
        return "branch_campus"
    if "closure of " in lower:
        return "main_campus"
    institution_teachout_markers = [
        "teach-out agreement",
        "teach-out agreements",
        "teach out agreement",
        "teach out agreements",
        "teach-out arrangement",
        "teach-out arrangements",
        "teach out arrangement",
        "teach out arrangements",
        "teach-out plan",
        "teach out plan",
    ]
    if any(marker in lower for marker in institution_teachout_markers) and not has_location_marker:
        return "main_campus"

    return None


def read_accreditation_closure_actions():
    rows = []
    if not ACCREDITATION_ACTIONS_CSV.exists():
        return rows

    for row in read_csv_rows(ACCREDITATION_ACTIONS_CSV):
        action_text = normalize_text(row.get("action_label_raw"))
        classification = classify_accreditation_action(action_text)
        if not classification:
            continue

        school_name = normalize_text(row.get("tracker_name") or row.get("institution_name_raw") or row.get("institution_name"))
        event_year = safe_int(row.get("action_year"))
        if event_year is None:
            continue

        notes = "Accreditation action references institution-level closure/teach-out or branch/main-campus teach-out."
        if "re-open" in action_text.lower() or "reopen" in action_text.lower():
            notes += " Action also mentions reopening at a new location."

        rows.append({
            "source_group": "accreditation_actions",
            "source_file": ACCREDITATION_ACTIONS_CSV.name,
            "source_section": normalize_text(row.get("accreditor")),
            "event_label": "Accreditation closure / teach-out action",
            "school_name_source": school_name,
            "state": normalize_text(row.get("tracker_state") or row.get("institution_state_raw") or row.get("state")),
            "opeid8": "",
            "opeid6": "",
            "unitid": normalize_text(row.get("unitid")),
            "event_date": normalize_text(row.get("action_date")),
            "event_year": event_year,
            "record_kind": "closure" if classification in {"main_campus", "branch_campus"} else "merger_or_consolidation",
            "classification": classification,
            "classification_basis": "Accreditation action about institution-wide closure/teach-out or branch/main-campus teach-out (program-only teach-outs excluded).",
            "sch_type": None,
            "closed_count": None,
            "pgm_length": None,
            "source_title": normalize_text(row.get("source_title")),
            "source_url": normalize_text(row.get("source_page_url") or row.get("source_url")),
            "notes": f"{notes} {action_text}".strip(),
        })

    return rows


def build_merger_candidates(closing_rows, federal_keys_by_unitid, federal_name_state_keys, opening_rows, status_index):
    opening_by_state = defaultdict(list)
    for row in opening_rows:
        opening_by_state[row["state"]].append(row["school_name_source"])

    candidates = []
    for row in closing_rows:
        if row["unitid"] in federal_keys_by_unitid:
            continue
        if slug_key(row["state"], row["school_name_source"]) in federal_name_state_keys:
            continue
        notes = "IPEDS closing/combining/leaving list entry not matched to a federal closure row."
        if opening_by_state.get(row["state"]):
            notes += f" Same-state openings also appear in the September 2024 IPEDS change workbook."
        unitid_status = status_index.get(row["unitid"], {})
        last_act = normalize_text(unitid_status.get("last_act"))
        classification = "main_campus" if last_act == "D" else "merger_consolidation"
        basis = (
            "IPEDS closing/combining/leaving list without a matched federal closure row; annual IPEDS status suggests the institution was marked closed."
            if classification == "main_campus"
            else "IPEDS closing/combining/leaving list without a matched federal closure row."
        )
        candidates.append({
            "source_group": "ipeds_change_workbook",
            "source_file": row["source_file"],
            "source_section": row["source_section"],
            "event_label": "IPEDS exit / likely merger or consolidation",
            "school_name_source": row["school_name_source"],
            "state": row["state"],
            "opeid8": "",
            "opeid6": "",
            "unitid": row["unitid"],
            "event_date": "",
            "event_year": row["event_year"],
            "record_kind": "closure" if classification == "main_campus" else "merger_or_consolidation",
            "classification": classification,
            "classification_basis": basis,
            "sch_type": None,
            "closed_count": None,
            "pgm_length": None,
            "source_title": "",
            "source_url": "",
            "notes": notes,
        })
    return candidates


def derive_ipeds_exit_candidates(status_index, federal_keys_by_unitid, federal_name_state_keys):
    # Extend the merger/consolidation review list beyond 2024 by scanning the
    # annual IPEDS directory for schools that disappear after carrying a
    # non-active final status.
    candidates = []
    for unitid, status in status_index.items():
        last_year = safe_int(status.get("last_year"))
        if last_year is None or last_year < MIN_EVENT_YEAR or last_year >= 2024:
            continue

        last_act = normalize_text(status.get("last_act"))
        if last_act not in {"C", "D"}:
            continue
        if unitid in federal_keys_by_unitid:
            continue
        if slug_key(status.get("last_state"), status.get("last_name")) in federal_name_state_keys:
            continue

        event_year = parse_closedat_year(status.get("last_closedat")) or (last_year + 1)
        if event_year < MIN_EVENT_YEAR or event_year > MAX_EVENT_YEAR:
            continue

        classification = "main_campus" if last_act == "D" else "merger_consolidation"
        event_label = "IPEDS annual status exit"
        basis = (
            "IPEDS annual directory status plus disappearance from later directories; last status marked the institution closed."
            if classification == "main_campus"
            else "IPEDS annual directory status plus disappearance from later directories."
        )
        candidates.append({
            "source_group": "ipeds_annual_directory",
            "source_file": f"HD{last_year}.zip",
            "source_section": "Annual HD directory",
            "event_label": event_label,
            "school_name_source": status.get("last_name"),
            "state": status.get("last_state"),
            "opeid8": status.get("last_opeid8"),
            "opeid6": normalize_text(status.get("last_opeid8"))[:6].lstrip("0") if status.get("last_opeid8") else "",
            "unitid": unitid,
            "event_date": status.get("last_closedat") if parse_closedat_year(status.get("last_closedat")) else "",
            "event_year": event_year,
            "record_kind": "closure" if classification == "main_campus" else "merger_or_consolidation",
            "classification": classification,
            "classification_basis": basis,
            "sch_type": None,
            "closed_count": None,
            "pgm_length": None,
            "source_title": "",
            "source_url": "",
            "notes": f"Last annual IPEDS status was ACT={last_act} in HD{last_year}.",
        })
    return candidates


def combine_monthly_records(section1_records, section2_records):
    grouped = {}
    for record in section1_records + section2_records:
        key = slug_key(record.get("opeid8"), record.get("event_date"), record.get("school_name_source"))
        if key not in grouped:
            grouped[key] = dict(record)
            grouped[key]["source_section"] = normalize_text(record.get("source_section"))
        else:
            sections = {
                part for part in [grouped[key].get("source_section"), record.get("source_section")] if part
            }
            grouped[key]["source_section"] = " + ".join(sorted(sections))
    return list(grouped.values())


def dedupe_running_records(records):
    # Collapse exact duplicate closure events across source files while keeping
    # multiple source labels together on one running-list row.
    grouped = {}
    for record in records:
        key = slug_key(
            record.get("classification"),
            record.get("opeid8") or record.get("opeid6") or record.get("unitid"),
            record.get("event_date") or record.get("event_year"),
            record.get("school_name_source"),
        )
        if key not in grouped:
            grouped[key] = dict(record)
        else:
            source_files = {
                part for part in [grouped[key].get("source_file"), record.get("source_file")] if part
            }
            source_sections = {
                part for part in [grouped[key].get("source_section"), record.get("source_section")] if part
            }
            grouped[key]["source_file"] = " + ".join(sorted(source_files))
            grouped[key]["source_section"] = " + ".join(sorted(source_sections))
    return list(grouped.values())


def sort_records(records):
    return sorted(
        records,
        key=lambda row: (
            safe_int(row.get("event_year")) or 9999,
            normalize_text(row.get("event_date")) or "9999-99-99",
            normalize_text(row.get("classification")),
            normalize_text(row.get("institution_name_ipeds") or row.get("school_name_source")),
        )
    )


def write_csv(path, rows, fieldnames):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_tracker_lookup(main_campus_rows):
    lookup = {}
    match_rows = []
    federal_source_groups = {"baseline_closure_file", "peps_closures_2024_sheet", "monthly_closed_school_report"}
    for row in main_campus_rows:
        if normalize_text(row.get("source_group")) not in federal_source_groups:
            continue
        unitid = normalize_text(row.get("unitid"))
        if not unitid:
            continue
        lookup[unitid] = {
            "unitid": unitid,
            "institution_name": row.get("institution_name_ipeds") or row.get("school_name_source"),
            "state": row.get("state"),
            "close_year": row.get("event_year"),
            "closure_source_date": "2025-04-03",
            "federal_school_name": row.get("school_name_source"),
            "opeid6": row.get("opeid6"),
        }
        match_rows.append(lookup[unitid])

    (REPO_ROOT / "data").mkdir(parents=True, exist_ok=True)
    with (REPO_ROOT / "data" / "closure_status_by_unitid.json").open("w", encoding="utf-8") as handle:
        json.dump(
            {
                "source_file": BASELINE_CLOSURE_XLSX.name,
                "as_of_date": "2025-04-03",
                "schools": lookup,
            },
            handle,
            indent=2,
            ensure_ascii=False,
        )

    write_csv(
        REPO_ROOT / "federal_closure" / "closure_status_tracker_matches.csv",
        sort_records(match_rows),
        ["unitid", "institution_name", "state", "close_year", "closure_source_date", "federal_school_name", "opeid6"],
    )


def main():
    rows_by_unitid, opeid8_to_unitid, opeid6_to_unitid, annual_hd_by_year, annual_effy_by_year, annual_hd_rows_by_unitid = load_ipeds_lookups()
    ipeds_status_index = build_ipeds_status_index(annual_hd_rows_by_unitid)

    baseline_main = read_baseline_closure_file()
    peps_rows, ipeds_closing_rows, ipeds_opening_rows = read_sep24_workbook()
    monthly_section_1 = parse_monthly_pdf_zip(MONTHLY_SECTION_1_ZIP, "Section I")
    monthly_section_2 = parse_monthly_pdf_zip(MONTHLY_SECTION_2_ZIP, "Section II")
    monthly_records = combine_monthly_records(monthly_section_1, monthly_section_2)
    accreditation_records = read_accreditation_closure_actions()

    # Build a federal closure key set so we can keep IPEDS-only exits out of
    # the outright closure tabs and reserve them for merger/consolidation review.
    federal_keys_by_unitid = set()

    enriched_baseline = [enrich_record(row, rows_by_unitid, opeid8_to_unitid, opeid6_to_unitid, annual_hd_by_year, annual_effy_by_year, ipeds_status_index) for row in baseline_main]
    for row in enriched_baseline:
        if row.get("unitid"):
            federal_keys_by_unitid.add(row["unitid"])

    enriched_peps = [enrich_record(row, rows_by_unitid, opeid8_to_unitid, opeid6_to_unitid, annual_hd_by_year, annual_effy_by_year, ipeds_status_index) for row in peps_rows]
    filtered_peps = []
    baseline_key_set = {
        slug_key(row.get("opeid6"), row.get("event_year"), row.get("school_name_source"))
        for row in enriched_baseline
    }
    for row in enriched_peps:
        if row["classification"] == "main_campus":
            main_key = slug_key(row.get("opeid6"), row.get("event_year"), row.get("school_name_source"))
            if main_key in baseline_key_set:
                continue
        if row.get("unitid"):
            federal_keys_by_unitid.add(row["unitid"])
        filtered_peps.append(row)

    enriched_monthly = [enrich_record(row, rows_by_unitid, opeid8_to_unitid, opeid6_to_unitid, annual_hd_by_year, annual_effy_by_year, ipeds_status_index) for row in monthly_records]
    for row in enriched_monthly:
        if row.get("unitid"):
            federal_keys_by_unitid.add(row["unitid"])

    federal_name_state_keys = {
        slug_key(row.get("state"), row.get("school_name_source"))
        for row in (enriched_baseline + filtered_peps + enriched_monthly)
    }
    merger_candidates = build_merger_candidates(
        ipeds_closing_rows,
        federal_keys_by_unitid,
        federal_name_state_keys,
        ipeds_opening_rows,
        ipeds_status_index,
    )
    merger_candidates.extend(
        derive_ipeds_exit_candidates(
            ipeds_status_index,
            federal_keys_by_unitid,
            federal_name_state_keys,
        )
    )
    enriched_mergers = [enrich_record(row, rows_by_unitid, opeid8_to_unitid, opeid6_to_unitid, annual_hd_by_year, annual_effy_by_year, ipeds_status_index) for row in merger_candidates]
    enriched_accreditation = [enrich_record(row, rows_by_unitid, opeid8_to_unitid, opeid6_to_unitid, annual_hd_by_year, annual_effy_by_year, ipeds_status_index) for row in accreditation_records]

    running_records = dedupe_running_records(enriched_baseline + filtered_peps + enriched_monthly + enriched_mergers + enriched_accreditation)
    running_records = [
        row for row in running_records
        if safe_int(row.get("event_year")) is not None and MIN_EVENT_YEAR <= safe_int(row.get("event_year")) <= MAX_EVENT_YEAR
    ]
    running_records = sort_records(running_records)

    main_campus_rows = [row for row in running_records if row.get("classification") == "main_campus"]
    branch_rows = [row for row in running_records if row.get("classification") == "branch_campus"]
    merger_rows = [row for row in running_records if row.get("classification") == "merger_consolidation"]
    federal_source_groups = {"baseline_closure_file", "peps_closures_2024_sheet", "monthly_closed_school_report"}
    private_sector_federal_main_rows = [
        row for row in main_campus_rows
        if row.get("source_group") in federal_source_groups
        and row.get("sector") in {"Private not-for-profit", "Private for-profit"}
    ]

    fieldnames = [
        "event_year",
        "event_date",
        "classification",
        "record_kind",
        "school_name_source",
        "institution_name_ipeds",
        "unitid",
        "unitid_match_method",
        "opeid8",
        "opeid6",
        "state",
        "sector",
        "institutional_category",
        "tracker_category",
        "student_headcount",
        "student_headcount_year",
        "ipeds_first_seen_year",
        "ipeds_last_seen_year",
        "ipeds_last_act",
        "ipeds_last_closedat",
        "institution_status",
        "is_active",
        "source_group",
        "source_file",
        "source_section",
        "event_label",
        "source_title",
        "source_url",
        "classification_basis",
        "sch_type",
        "closed_count",
        "pgm_length",
        "notes",
    ]

    write_csv(RUNNING_OUTPUT, running_records, fieldnames)
    write_csv(MAIN_OUTPUT, sort_records(main_campus_rows), fieldnames)
    write_csv(BRANCH_OUTPUT, sort_records(branch_rows), fieldnames)
    write_csv(MERGER_OUTPUT, sort_records(merger_rows), fieldnames)
    write_csv(PRIVATE_FEDERAL_MAIN_OUTPUT, sort_records(private_sector_federal_main_rows), fieldnames)

    summary = {
        "baseline_main_rows": len(enriched_baseline),
        "peps_rows_kept": len(filtered_peps),
        "monthly_rows_kept": len(enriched_monthly),
        "merger_candidates": len(enriched_mergers),
        "running_rows": len(running_records),
        "main_campus_rows": len(main_campus_rows),
        "branch_rows": len(branch_rows),
        "merger_rows": len(merger_rows),
        "private_sector_federal_main_rows": len(private_sector_federal_main_rows),
        "accreditation_rows_kept": len(enriched_accreditation),
        "min_event_year": MIN_EVENT_YEAR,
        "max_event_year": MAX_EVENT_YEAR,
        "studentaid_confirmation_note": (
            "StudentAid closed-school pages are treated as an official confirmation layer for borrower-facing closure status, "
            "but not as the machine-readable backbone for this dataset."
        ),
    }

    with SUMMARY_OUTPUT.open("w", encoding="utf-8") as handle:
        json.dump(summary, handle, indent=2)

    # Refresh the tracker lookup from the main-campus closure tab so the school
    # flag is consistent with the workbook outputs.
    build_tracker_lookup(main_campus_rows)

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
