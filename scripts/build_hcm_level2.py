import csv
import json
from collections import defaultdict
from datetime import date
from pathlib import Path

import xlrd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_PIPELINES = ROOT / "data_pipelines"

# Federal Student Aid publishes HCM files quarterly. This script is intentionally
# source-versioned to the raw snapshot workbooks listed below, not to a generic
# calendar year. Add a new SOURCE_FILES entry and update the output/version label
# when a newer quarterly workbook is adopted.
SOURCE_VERSION_LABEL = "December 2024 through December 2025 HCM quarterly snapshots"
SOURCE_FILES = [
    {
        "snapshot_date": "2024-12-01",
        "snapshot_label": "December 2024",
        "path": DATA_PIPELINES / "federal_hcm" / "raw" / "schools-on-hcm-dec-2024.xls",
        "kind": "xls",
    },
    {
        "snapshot_date": "2025-03-01",
        "snapshot_label": "March 2025",
        "path": DATA_PIPELINES / "federal_hcm" / "raw" / "schools-on-hcm-mar-2025.xlsx",
        "kind": "xlsx",
    },
    {
        "snapshot_date": "2025-06-01",
        "snapshot_label": "June 2025",
        "path": DATA_PIPELINES / "federal_hcm" / "raw" / "schools-on-hcm-jun-2025.xlsx",
        "kind": "xlsx",
    },
    {
        "snapshot_date": "2025-09-01",
        "snapshot_label": "September 2025",
        "path": DATA_PIPELINES / "federal_hcm" / "raw" / "schools-on-hcm-sep-2025.xlsx",
        "kind": "xlsx",
    },
    {
        "snapshot_date": "2025-12-01",
        "snapshot_label": "December 2025",
        "path": DATA_PIPELINES / "federal_hcm" / "raw" / "schools-on-hcm-dec-2025.xlsx",
        "kind": "xlsx",
    },
]

IPEDS_RAW = ROOT / "ipeds" / "raw" / "ipeds_financial_health_raw_2014_2024.csv"
TRACKER_INDEX = ROOT / "data" / "schools_index.json"

OUTPUT_DIR = DATA_PIPELINES / "federal_hcm"
OUTPUT_ALL = OUTPUT_DIR / "hcm_level2_snapshots_2024_2025.csv"
OUTPUT_SUMMARY = OUTPUT_DIR / "hcm_level2_summary.csv"
OUTPUT_DEC24_DROP = OUTPUT_DIR / "hcm2_dec2024_dropped_since.csv"
OUTPUT_MAR25_DROP = OUTPUT_DIR / "hcm2_mar2025_dropped_since.csv"
OUTPUT_JUN25_DROP = OUTPUT_DIR / "hcm2_jun2025_dropped_since.csv"
OUTPUT_DEC24_STAY = OUTPUT_DIR / "hcm2_dec2024_remained_since.csv"
OUTPUT_MAR25_STAY = OUTPUT_DIR / "hcm2_mar2025_remained_since.csv"
OUTPUT_HCM1 = OUTPUT_DIR / "hcm_level1_snapshots_2024_2025.csv"
OUTPUT_DOWNGRADE = OUTPUT_DIR / "hcm2_to_hcm1_downgrades_2025.csv"
OUTPUT_JSON = ROOT / "data" / "hcm2_by_unitid.json"


def ensure_required_inputs(script_name, requirements):
    missing = []
    for requirement in requirements:
        if requirement["path"].exists():
            continue
        missing.append(requirement)

    if not missing:
        return

    message_lines = [
        f"{script_name} cannot start because required local input files are missing.",
        "",
        "Add the missing files below and then run the script again.",
        "",
    ]

    for requirement in missing:
        message_lines.append(f"- {requirement['label']}")
        message_lines.append(f"  Expected path: {requirement['path']}")
        if requirement.get("note"):
            message_lines.append(f"  Why it is needed: {requirement['note']}")
        message_lines.append("")

    raise FileNotFoundError("\n".join(message_lines).rstrip())


def normalize_opeid(value):
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    return digits.zfill(8) if digits else ""


def map_sector_label(hcm_type):
    text = str(hcm_type or "").strip()
    mapping = {
        "Public": "Public",
        "Private Non-Profit": "Private not-for-profit",
        "Foreign Private Non-Profit": "Private not-for-profit",
        "Proprietary": "Private for-profit",
        "Foreign For-Profit": "Private for-profit",
    }
    return mapping.get(text, text)


def build_ipeds_lookup():
    latest_by_opeid = {}
    with IPEDS_RAW.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            unitid = str(row.get("unitid", "")).strip()
            opeid = normalize_opeid(row.get("opeid", ""))
            year = str(row.get("year", "")).strip()
            if not unitid or not opeid:
                continue
            current = latest_by_opeid.get(opeid)
            if current is None or year > current["year"]:
                latest_by_opeid[opeid] = {
                    "year": year,
                    "unitid": unitid,
                    "institution_name": row.get("institution_name", ""),
                    "city": row.get("city", ""),
                    "state": row.get("state", ""),
                }
    return latest_by_opeid


def build_tracker_unitid_set():
    with TRACKER_INDEX.open(encoding="utf-8") as handle:
        rows = json.load(handle)
    return {str(row.get("unitid", "")).strip() for row in rows if row.get("unitid")}


def parse_hcm_rows(snapshot, level_number):
    level_tag = f"HCM {level_number}"
    if snapshot["kind"] == "xls":
        book = xlrd.open_workbook(snapshot["path"])
        sheet_name = next(name for name in book.sheet_names() if level_tag in name)
        sheet = book.sheet_by_name(sheet_name)
        headers = [str(value).strip() for value in sheet.row_values(2)]
        rows = []
        for row_idx in range(3, sheet.nrows):
            values = sheet.row_values(row_idx)
            if not any(str(value).strip() for value in values):
                continue
            rows.append(dict(zip(headers, values)))
        return rows

    workbook = load_workbook(snapshot["path"], read_only=True, data_only=True)
    sheet_name = next(name for name in workbook.sheetnames if level_tag in name)
    sheet = workbook[sheet_name]
    row_iter = sheet.iter_rows(values_only=True)
    next(row_iter)
    next(row_iter)
    headers = [str(value).strip() if value is not None else "" for value in next(row_iter)]
    rows = []
    for values in row_iter:
        if not values or not any(str(value).strip() for value in values if value is not None):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def normalize_snapshot_rows(ipeds_by_opeid, tracker_unitids, level_number):
    all_rows = []
    for snapshot in SOURCE_FILES:
        raw_rows = parse_hcm_rows(snapshot, level_number)
        deduped = {}
        for row in raw_rows:
            opeid = normalize_opeid(row.get("OPE ID", ""))
            if not opeid:
                continue
            deduped[opeid] = row

        for opeid, row in deduped.items():
            ipeds_match = ipeds_by_opeid.get(opeid, {})
            institution_type = (
                row.get("Institution Type Desc")
                or row.get("Institution Type")
                or ""
            )
            unitid = str(ipeds_match.get("unitid", "")).strip()
            all_rows.append(
                {
                    "snapshot_date": snapshot["snapshot_date"],
                    "snapshot_label": snapshot["snapshot_label"],
                    "hcm_level": f"HCM{level_number}",
                    "opeid": opeid,
                    "unitid": unitid,
                    "institution_name": row.get("Institution Name") or ipeds_match.get("institution_name", ""),
                    "city": row.get("City") or ipeds_match.get("city", ""),
                    "state": row.get("State") or ipeds_match.get("state", ""),
                    "country": row.get("Country", ""),
                    "sector": map_sector_label(institution_type),
                    "institution_type_desc": institution_type,
                    "reason_on_description": row.get("Reason On Description")
                    or row.get("Method Reason Desc")
                    or "",
                    "in_finance_tracker": "Yes" if unitid and unitid in tracker_unitids else "No",
                }
            )

    return sorted(all_rows, key=lambda row: (row["snapshot_date"], row["state"], row["institution_name"]))


def first_missing_snapshot(after_index, present_indices):
    for idx in range(after_index + 1, len(SOURCE_FILES)):
        if idx not in present_indices:
            return SOURCE_FILES[idx]["snapshot_label"]
    return None


def build_histories(all_rows):
    order = [item["snapshot_date"] for item in SOURCE_FILES]
    label_by_date = {item["snapshot_date"]: item["snapshot_label"] for item in SOURCE_FILES}

    rows_by_opeid = defaultdict(list)
    for row in all_rows:
        rows_by_opeid[row["opeid"]].append(row)

    histories = {}
    for opeid, rows in rows_by_opeid.items():
        rows = sorted(rows, key=lambda row: order.index(row["snapshot_date"]))
        dates = [row["snapshot_date"] for row in rows]
        present_indices = {order.index(snapshot_date) for snapshot_date in dates}
        last_index = max(present_indices)
        first_off_label = first_missing_snapshot(last_index, present_indices)
        histories[opeid] = {
            "opeid": opeid,
            "institution_name": rows[-1]["institution_name"],
            "unitid": rows[-1]["unitid"],
            "state": rows[-1]["state"],
            "sector": rows[-1]["sector"],
            "in_finance_tracker": rows[-1]["in_finance_tracker"],
            "snapshots_present": dates,
            "snapshot_labels_present": [label_by_date[d] for d in dates],
            "first_snapshot_date": dates[0],
            "first_snapshot_label": label_by_date[dates[0]],
            "latest_snapshot_date_present": dates[-1],
            "latest_snapshot_label_present": label_by_date[dates[-1]],
            "on_latest_snapshot": dates[-1] == order[-1],
            "first_snapshot_absent_after_last_presence": first_off_label or "",
            "latest_reason_on_description": rows[-1]["reason_on_description"],
        }
    return histories


def build_downgrade_rows(hcm1_histories, hcm2_histories):
    order = [item["snapshot_date"] for item in SOURCE_FILES]
    label_by_date = {item["snapshot_date"]: item["snapshot_label"] for item in SOURCE_FILES}

    rows = []
    for opeid, history in hcm2_histories.items():
        hcm2_present = set(history["snapshots_present"])
        hcm1_history = hcm1_histories.get(opeid)
        if not hcm1_history:
            continue

        hcm1_present = set(hcm1_history["snapshots_present"])
        last_hcm2_date = history["latest_snapshot_date_present"]
        last_hcm2_index = order.index(last_hcm2_date)

        first_hcm1_after = None
        for snapshot_date in order[last_hcm2_index + 1 :]:
            if snapshot_date in hcm1_present:
                first_hcm1_after = snapshot_date
                break

        if not first_hcm1_after:
            continue

        rows.append(
            {
                "institution_name": history["institution_name"],
                "sector": history["sector"],
                "state": history["state"],
                "unitid": history["unitid"],
                "opeid": opeid,
                "last_hcm2_snapshot_label": label_by_date[last_hcm2_date],
                "first_hcm1_snapshot_after_hcm2_label": label_by_date[first_hcm1_after],
                "latest_hcm2_reason_on_description": history["latest_reason_on_description"],
                "in_finance_tracker": history["in_finance_tracker"],
            }
        )

    return sorted(rows, key=lambda row: (row["state"], row["institution_name"]))


def make_transition_rows(histories, start_snapshot_date, require_all_future):
    order = [item["snapshot_date"] for item in SOURCE_FILES]
    start_index = order.index(start_snapshot_date)
    future_dates = order[start_index + 1 :]

    rows = []
    for history in histories.values():
        present = set(history["snapshots_present"])
        if start_snapshot_date not in present:
            continue

        if require_all_future:
            keep = all(snapshot_date in present for snapshot_date in future_dates)
        else:
            keep = SOURCE_FILES[-1]["snapshot_date"] not in present

        if not keep:
            continue

        rows.append(
            {
                "institution_name": history["institution_name"],
                "sector": history["sector"],
                "state": history["state"],
                "unitid": history["unitid"],
                "opeid": history["opeid"],
                "first_snapshot_label": history["first_snapshot_label"],
                "latest_snapshot_label_present": history["latest_snapshot_label_present"],
                "first_snapshot_absent_after_last_presence": history["first_snapshot_absent_after_last_presence"],
                "latest_reason_on_description": history["latest_reason_on_description"],
                "in_finance_tracker": history["in_finance_tracker"],
            }
        )

    return sorted(rows, key=lambda row: (row["state"], row["institution_name"]))


def build_school_lookup(histories):
    order = [item["snapshot_date"] for item in SOURCE_FILES]
    label_by_date = {item["snapshot_date"]: item["snapshot_label"] for item in SOURCE_FILES}

    grouped = defaultdict(list)
    for history in histories.values():
        unitid = history["unitid"]
        if not unitid:
            continue
        grouped[unitid].append(history)

    schools = {}
    for unitid, items in grouped.items():
        snapshot_dates = sorted(
            {snapshot for item in items for snapshot in item["snapshots_present"]},
            key=order.index,
        )
        first_snapshot = snapshot_dates[0]
        latest_present = snapshot_dates[-1]
        on_latest = latest_present == order[-1]
        last_index = order.index(latest_present)
        first_off_label = None
        for idx in range(last_index + 1, len(order)):
            if order[idx] not in snapshot_dates:
                first_off_label = label_by_date[order[idx]]
                break

        schools[unitid] = {
            "unitid": unitid,
            "institution_name": items[0]["institution_name"],
            "sector": items[0]["sector"],
            "state": items[0]["state"],
            "opeids": sorted({item["opeid"] for item in items}),
            "snapshots_present": snapshot_dates,
            "snapshot_labels_present": [label_by_date[d] for d in snapshot_dates],
            "first_snapshot_label": label_by_date[first_snapshot],
            "latest_snapshot_label_present": label_by_date[latest_present],
            "on_latest_snapshot": on_latest,
            "first_snapshot_absent_after_last_presence": first_off_label or "",
            "latest_reason_on_description": items[-1]["latest_reason_on_description"],
        }

    return schools


def apply_downgrade_lookup(schools, downgrade_rows):
    by_unitid = defaultdict(list)
    for row in downgrade_rows:
        unitid = str(row.get("unitid", "")).strip()
        if not unitid:
            continue
        by_unitid[unitid].append(row)

    for unitid, items in by_unitid.items():
        items = sorted(
            items,
            key=lambda row: [item["snapshot_label"] for item in SOURCE_FILES].index(row["first_hcm1_snapshot_after_hcm2_label"])
            if row["first_hcm1_snapshot_after_hcm2_label"] in [item["snapshot_label"] for item in SOURCE_FILES]
            else 999
        )
        schools.setdefault(unitid, {"unitid": unitid})
        schools[unitid]["downgraded_to_hcm1_after_hcm2"] = True
        schools[unitid]["first_hcm2_snapshot_before_downgrade_label"] = items[0]["last_hcm2_snapshot_label"]
        schools[unitid]["first_hcm1_snapshot_after_hcm2_label"] = items[0]["first_hcm1_snapshot_after_hcm2_label"]
        schools[unitid]["latest_hcm2_reason_on_description"] = items[0]["latest_hcm2_reason_on_description"]

    for school in schools.values():
        school.setdefault("downgraded_to_hcm1_after_hcm2", False)
        school.setdefault("first_hcm2_snapshot_before_downgrade_label", "")
        school.setdefault("first_hcm1_snapshot_after_hcm2_label", "")

    return schools


def write_csv(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        with path.open("w", newline="", encoding="utf-8") as handle:
            handle.write("")
        return

    fieldnames = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main():
    ensure_required_inputs(
        "build_hcm_level2.py",
        [
            *[
                {
                    "label": f"HCM raw workbook for {snapshot['snapshot_label']}",
                    "path": snapshot["path"],
                    "note": "Used to build the quarterly HCM history shown on school profiles.",
                }
                for snapshot in SOURCE_FILES
            ],
            {
                "label": "IPEDS raw dataset",
                "path": IPEDS_RAW,
                "note": "Used to map federal OPEIDs back to UNITIDs and school names.",
            },
            {
                "label": "Schools index JSON",
                "path": TRACKER_INDEX,
                "note": "Used to flag which HCM schools are in the interactive tracker universe.",
            },
        ],
    )
    ipeds_by_opeid = build_ipeds_lookup()
    tracker_unitids = build_tracker_unitid_set()
    hcm2_rows = normalize_snapshot_rows(ipeds_by_opeid, tracker_unitids, level_number=2)
    hcm1_rows = normalize_snapshot_rows(ipeds_by_opeid, tracker_unitids, level_number=1)
    histories = build_histories(hcm2_rows)
    hcm1_histories = build_histories(hcm1_rows)
    downgrade_rows = build_downgrade_rows(hcm1_histories, histories)
    school_lookup = build_school_lookup(histories)
    school_lookup = apply_downgrade_lookup(school_lookup, downgrade_rows)

    dec24_drop = make_transition_rows(histories, "2024-12-01", require_all_future=False)
    mar25_drop = make_transition_rows(histories, "2025-03-01", require_all_future=False)
    jun25_drop = make_transition_rows(histories, "2025-06-01", require_all_future=False)
    dec24_stay = make_transition_rows(histories, "2024-12-01", require_all_future=True)
    mar25_stay = make_transition_rows(histories, "2025-03-01", require_all_future=True)

    counts_by_snapshot = defaultdict(int)
    tracker_counts_by_snapshot = defaultdict(int)
    for row in hcm2_rows:
        counts_by_snapshot[row["snapshot_label"]] += 1
        if row["in_finance_tracker"] == "Yes":
            tracker_counts_by_snapshot[row["snapshot_label"]] += 1

    summary_rows = []
    for snapshot in SOURCE_FILES:
        label = snapshot["snapshot_label"]
        summary_rows.append(
            {
                "snapshot_date": snapshot["snapshot_date"],
                "snapshot_label": label,
                "hcm2_institutions": counts_by_snapshot[label],
                "hcm2_institutions_in_finance_tracker": tracker_counts_by_snapshot[label],
            }
        )

    write_csv(OUTPUT_ALL, hcm2_rows)
    write_csv(OUTPUT_HCM1, hcm1_rows)
    write_csv(OUTPUT_SUMMARY, summary_rows)
    write_csv(OUTPUT_DEC24_DROP, dec24_drop)
    write_csv(OUTPUT_MAR25_DROP, mar25_drop)
    write_csv(OUTPUT_JUN25_DROP, jun25_drop)
    write_csv(OUTPUT_DEC24_STAY, dec24_stay)
    write_csv(OUTPUT_MAR25_STAY, mar25_stay)
    write_csv(OUTPUT_DOWNGRADE, downgrade_rows)

    first_label = SOURCE_FILES[0]["snapshot_label"]
    latest_label = SOURCE_FILES[-1]["snapshot_label"]
    trump_drop_from = counts_by_snapshot[first_label]
    trump_drop_to = counts_by_snapshot[latest_label]

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_JSON.open("w", encoding="utf-8") as handle:
        json.dump(
            {
                "generated_at": date.today().isoformat(),
                "source_version_label": SOURCE_VERSION_LABEL,
                "source_files": [str(item["path"].relative_to(ROOT)).replace("\\", "/") for item in SOURCE_FILES],
                "summary": {
                    "counts_by_snapshot": summary_rows,
                    "trump_administration_drop_from": trump_drop_from,
                    "trump_administration_drop_to": trump_drop_to,
                    "latest_snapshot_label": latest_label,
                    "dec2024_drop_count": len(dec24_drop),
                    "mar2025_drop_count": len(mar25_drop),
                    "jun2025_drop_count": len(jun25_drop),
                    "dec2024_remained_count": len(dec24_stay),
                    "mar2025_remained_count": len(mar25_stay),
                    "downgraded_to_hcm1_count": len(downgrade_rows),
                },
                "schools": school_lookup,
            },
            handle,
            indent=2,
        )

    print(f"Wrote {len(hcm2_rows)} HCM2 snapshot rows to {OUTPUT_ALL}")
    print(f"Wrote {len(hcm1_rows)} HCM1 snapshot rows to {OUTPUT_HCM1}")
    print(f"Wrote {len(downgrade_rows)} HCM2-to-HCM1 downgrades to {OUTPUT_DOWNGRADE}")
    print(f"Wrote {len(school_lookup)} unitid-matched school histories to {OUTPUT_JSON}")
    print(
        f"HCM2 count fell from "
        f"{trump_drop_from} in {first_label} to {trump_drop_to} in {latest_label}."
    )


if __name__ == "__main__":
    try:
        main()
    except (FileNotFoundError, RuntimeError) as exc:
        raise SystemExit(str(exc))
