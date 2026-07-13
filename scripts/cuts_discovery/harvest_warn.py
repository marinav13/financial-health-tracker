#!/usr/bin/env python3
import io
import re
import urllib.parse
from datetime import datetime, timezone
from html import unescape
from pathlib import Path

import openpyxl

if __package__ in (None, ""):
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    from cuts_discovery.common import lead_id_for, parse_date_to_iso, today_iso
    from import_supabase_institution_mapping import STATE_ABBREV, normalize_name
else:
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    from .common import lead_id_for, parse_date_to_iso, today_iso
    from import_supabase_institution_mapping import STATE_ABBREV, normalize_name


CA_WARN_URL = "https://edd.ca.gov/siteassets/files/jobs_and_training/warn/warn_report.xlsx"
WA_WARN_URL = "https://fortress.wa.gov/esd/file/WARN"
NY_WARN_URL = "https://dol.ny.gov/warn-notices"
NY_BASE_URL = "https://dol.ny.gov"
WARNTRACKER_URL_TEMPLATE = "https://www.warntracker.com/?company={query}&tab=by-company-and-state"
WARNTRACKER_QUERIES = ("College", "University", "Institute", "Seminary")
WARN_LOOKBACK_DAYS = 180


def _trim(value) -> str:
    return str(value or "").strip()


def _clean_text(value) -> str:
    return re.sub(r"\s+", " ", unescape(_trim(value))).strip()


def _count_text(value) -> str:
    count = _trim(value).replace(",", "")
    if not count:
        return "unspecified employees"
    return f"{count} employees"


def _iso_date(value) -> str:
    if hasattr(value, "date"):
        return value.date().isoformat()
    return parse_date_to_iso(_trim(value))


def _age_days(iso_date: str) -> float | None:
    if not iso_date:
        return None
    return (datetime.now(timezone.utc).date() - datetime.fromisoformat(iso_date).date()).days


def _is_recent(iso_date: str, max_age_days: int = WARN_LOOKBACK_DAYS) -> bool:
    age_days = _age_days(iso_date)
    return age_days is not None and age_days <= max_age_days


def _warn_if_stale(source_name: str, latest_iso_date: str, expected_days: int = WARN_LOOKBACK_DAYS) -> None:
    age_days = _age_days(latest_iso_date)
    if age_days is None or age_days <= expected_days:
        return
    print(
        f"::warning::cuts discovery WARN source {source_name} appears stale: latest notice is "
        f"{latest_iso_date} ({age_days} days old)."
    )


def _matching_institution(employer: str, state_full: str, mapping_rows: list[dict]) -> dict | None:
    norm_employer = normalize_name(employer)
    if not norm_employer:
        return None

    matches = []
    padded_employer = f" {norm_employer} "
    for row in mapping_rows:
        norm_name = _trim(row.get("norm_name"))
        if not norm_name:
            continue
        if f" {norm_name} " in padded_employer:
            matches.append(row)

    if not matches:
        return None

    if state_full:
        state_matches = [
            row for row in matches if _trim(row.get("state")).lower() == _trim(state_full).lower()
        ]
        if len({row["unitid"] for row in state_matches}) == 1 and state_matches:
            return state_matches[0]

    if len({row["unitid"] for row in matches}) == 1:
        return matches[0]
    return None


def _event_kind(text: str) -> str:
    lower_text = _trim(text).lower()
    if "closure" in lower_text or "closing" in lower_text:
        return "closure"
    return "layoff"


def _build_warn_lead(
    *,
    employer: str,
    institution_row: dict,
    source_url: str,
    source_tag: str,
    source_publication: str,
    state_full: str,
    notice_date: str,
    count_value,
    event_text: str,
    effective_date: str = "",
    location_text: str = "",
    extra_text: str = "",
) -> dict:
    institution_name = _trim(institution_row.get("institution_name")) or employer
    event_kind = _event_kind(event_text)
    scale_text = _count_text(count_value)
    headline = f"{institution_name} WARN {event_kind} notice affecting {scale_text}"

    snippet_parts = [
        f"{state_full} WARN notice for {employer}.",
        f"Affects {scale_text}.",
        f"Type: {_clean_text(event_text)}.",
    ]
    if effective_date:
        snippet_parts.append(f"Effective date: {effective_date}.")
    if location_text:
        snippet_parts.append(f"Location: {location_text}.")
    if extra_text:
        snippet_parts.append(_clean_text(extra_text))

    normalized_url = _trim(source_url)
    return {
        "lead_id": lead_id_for(normalized_url),
        "first_seen": today_iso(),
        "tier": "warn",
        "query_or_feed": source_tag,
        "url": normalized_url,
        "publisher": source_publication,
        "headline": headline,
        "published_date": notice_date,
        "snippet": " ".join(part for part in snippet_parts if part)[:500],
        "status": "new",
        "status_reason": "",
    }


def parse_ca_rows(body: bytes) -> list[dict]:
    workbook = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    sheet = next(
        (workbook[name] for name in workbook.sheetnames if name.strip().startswith("Detailed WARN Report")),
        None,
    )
    if sheet is None:
        raise ValueError("California WARN workbook is missing the detailed report sheet.")

    rows = []
    latest_notice_date = ""
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 8:
            continue
        notice_date = _iso_date(row[1])
        latest_notice_date = max(latest_notice_date, notice_date)
        rows.append(
            {
                "county": _clean_text(row[0]),
                "notice_date": notice_date,
                "effective_date": _iso_date(row[3]),
                "employer": _clean_text(row[4]),
                "event_text": _clean_text(row[5]),
                "employee_count": _trim(row[6]),
                "address": _clean_text(row[7]),
            }
        )
    _warn_if_stale("ca", latest_notice_date)
    return rows


def harvest_ca_warn(fetcher, mapping_rows: list[dict]) -> list[dict]:
    rows = []
    for parsed in parse_ca_rows(fetcher.get(CA_WARN_URL, max_age_days=1.0)):
        if not _is_recent(parsed["notice_date"]):
            continue
        matched = _matching_institution(parsed["employer"], "California", mapping_rows)
        if matched is None:
            continue
        query = urllib.parse.urlencode(
            {
                "company": parsed["employer"],
                "county": parsed["county"],
                "notice_date": parsed["notice_date"],
            }
        )
        source_url = f"{CA_WARN_URL}?{query}"
        rows.append(
            _build_warn_lead(
                employer=parsed["employer"],
                institution_row=matched,
                source_url=source_url,
                source_tag="ca_warn",
                source_publication="California WARN",
                state_full="California",
                notice_date=parsed["notice_date"],
                count_value=parsed["employee_count"],
                event_text=parsed["event_text"],
                effective_date=parsed["effective_date"],
                location_text=parsed["address"],
            )
        )
    return rows


def parse_wa_rows(html: str) -> list[dict]:
    table_start = re.search(r'<table[^>]+id="ucPSW_gvMain"', html, flags=re.IGNORECASE)
    if not table_start:
        raise ValueError("Washington WARN page is missing the expected notice table.")
    table_html = html[table_start.start():]

    rows = []
    for cells_html in re.findall(r"<tr>([\s\S]*?)</tr>", table_html, flags=re.IGNORECASE):
        cell_values = re.findall(r"<td[^>]*>([\s\S]*?)</td>", cells_html, flags=re.IGNORECASE)
        if len(cell_values) < 8:
            continue
        notice_match = re.search(r"href='([^']+DownloadFile[^']+)'", cells_html, flags=re.IGNORECASE)
        if not notice_match:
            continue
        rows.append(
            {
                "employer": _clean_text(cell_values[0]),
                "location": _clean_text(cell_values[1]),
                "effective_date": parse_date_to_iso(_clean_text(cell_values[2])),
                "employee_count": _clean_text(cell_values[3]),
                "event_text": " ".join(
                    part for part in [_clean_text(cell_values[4]), _clean_text(cell_values[5])] if part
                ),
                "notice_date": parse_date_to_iso(_clean_text(cell_values[6])),
                "notice_href": urllib.parse.urljoin(WA_WARN_URL + "/", _clean_text(notice_match.group(1))),
            }
        )
    if not rows:
        raise ValueError("Washington WARN page parsed, but no notice rows were found.")
    return rows


def harvest_wa_warn(fetcher, mapping_rows: list[dict]) -> list[dict]:
    html = fetcher.get(WA_WARN_URL, max_age_days=0.9).decode("utf-8", "replace")
    parsed_rows = parse_wa_rows(html)
    _warn_if_stale("wa", max((row["notice_date"] for row in parsed_rows if row["notice_date"]), default=""))

    rows = []
    for parsed in parsed_rows:
        if not _is_recent(parsed["notice_date"]):
            continue
        matched = _matching_institution(parsed["employer"], "Washington", mapping_rows)
        if matched is None:
            continue
        rows.append(
            _build_warn_lead(
                employer=parsed["employer"],
                institution_row=matched,
                source_url=parsed["notice_href"],
                source_tag="wa_warn",
                source_publication="Washington WARN",
                state_full="Washington",
                notice_date=parsed["notice_date"],
                count_value=parsed["employee_count"],
                event_text=parsed["event_text"],
                effective_date=parsed["effective_date"],
                location_text=parsed["location"],
            )
        )
    return rows


def parse_ny_listing_rows(html: str) -> list[dict]:
    rows = []
    for row_html in re.findall(r"<tr[^>]*>([\s\S]*?)</tr>", html, flags=re.IGNORECASE):
        cell_values = re.findall(r"<td[^>]*>([\s\S]*?)</td>", row_html, flags=re.IGNORECASE)
        if len(cell_values) < 4:
            continue
        link_match = re.search(r'href="([^"]+/warn-[^"]+|/warn-[^"]+)"', row_html, flags=re.IGNORECASE)
        if not link_match:
            continue
        rows.append(
            {
                "detail_url": urllib.parse.urljoin(NY_BASE_URL, _clean_text(link_match.group(1))),
                "employer": _clean_text(re.sub(r"<[^>]+>", " ", cell_values[0])),
                "region": _clean_text(re.sub(r"<[^>]+>", " ", cell_values[1])),
                "date_posted": parse_date_to_iso(_clean_text(re.sub(r"<[^>]+>", " ", cell_values[2]))),
                "notice_date_text": _clean_text(re.sub(r"<[^>]+>", " ", cell_values[3])),
            }
        )
    if not rows:
        raise ValueError("New York WARN list page is missing the expected notice rows.")
    return rows


def extract_ny_pdf_text(body: bytes) -> str | None:
    if not body.lstrip().startswith(b"%PDF-"):
        return body.decode("utf-8", "replace")
    try:
        from pypdf import PdfReader
    except ModuleNotFoundError:
        return None
    try:
        reader = PdfReader(io.BytesIO(body))
        return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception:
        return body.decode("utf-8", "replace")


def parse_ny_pdf_fields(text: str) -> dict:
    flat = _clean_text(text)

    def match(pattern: str) -> str:
        found = re.search(pattern, flat, flags=re.IGNORECASE)
        return _clean_text(found.group(1)) if found else ""

    employer = match(r"Company:\s*(.*?)\s*(?:FEIN NUM:|Contact:)")
    affected = match(r"Total Number of Affected Workers:\s*([0-9,]+)")
    event_reason = match(r"Reason For Closure:\s*(.*?)\s*(?:Company:|Total Number of Affected Workers:)")
    notice_date = parse_date_to_iso(match(r"Date of Notice:\s*([A-Za-z0-9, /-]+)"))
    closure_end_date = parse_date_to_iso(match(r"Closure End Date:\s*([A-Za-z0-9, /-]+)"))
    site_count = match(r"Number of Affected Employees at Site:\s*([0-9,]+)")

    count_value = affected or site_count
    if not employer or not count_value:
        raise ValueError("New York WARN notice PDF is missing employer or affected-worker count.")

    return {
        "employer": employer,
        "employee_count": count_value,
        "event_text": (
            f"Closure {event_reason}".strip()
            if closure_end_date
            else f"Layoff {event_reason}".strip()
        ),
        "notice_date": notice_date,
        "effective_date": closure_end_date,
    }


def harvest_ny_warn(fetcher, mapping_rows: list[dict]) -> list[dict]:
    html = fetcher.get(NY_WARN_URL, max_age_days=0.9).decode("utf-8", "replace")
    parsed_rows = parse_ny_listing_rows(html)
    _warn_if_stale("ny", max((row["date_posted"] for row in parsed_rows if row["date_posted"]), default=""))

    rows = []
    for parsed in parsed_rows:
        if not _is_recent(parsed["date_posted"] or parse_date_to_iso(parsed["notice_date_text"])):
            continue
        matched = _matching_institution(parsed["employer"], "New York", mapping_rows)
        if matched is None:
            continue

        pdf_text = extract_ny_pdf_text(fetcher.get(parsed["detail_url"], max_age_days=6.0))
        if pdf_text is None:
            print("::warning::cuts discovery WARN source ny skipped: pypdf is not installed.")
            return rows
        pdf_fields = parse_ny_pdf_fields(pdf_text)
        employer = pdf_fields["employer"] or parsed["employer"]
        rows.append(
            _build_warn_lead(
                employer=employer,
                institution_row=matched,
                source_url=parsed["detail_url"],
                source_tag="ny_warn",
                source_publication="New York WARN",
                state_full="New York",
                notice_date=pdf_fields["notice_date"] or parsed["date_posted"],
                count_value=pdf_fields["employee_count"],
                event_text=pdf_fields["event_text"],
                effective_date=pdf_fields["effective_date"],
                location_text=parsed["region"],
            )
        )
    return rows


def parse_warntracker_rows(html: str) -> list[dict]:
    if "No records to display" in html:
        return []

    tbody_match = re.search(r"<tbody[^>]*>([\s\S]*?)</tbody>", html, flags=re.IGNORECASE)
    if not tbody_match:
        raise ValueError("WARNTracker page is missing the expected results table body.")

    rows = []
    for row_html in re.findall(r"<tr[^>]*>([\s\S]*?)</tr>", tbody_match.group(1), flags=re.IGNORECASE):
        cell_values = [
            _clean_text(re.sub(r"<[^>]+>", " ", value))
            for value in re.findall(r"<td[^>]*>([\s\S]*?)</td>", row_html, flags=re.IGNORECASE)
        ]
        if len(cell_values) < 7:
            continue
        if len(cell_values) >= 8:
            _, employer, state_code, employee_count, notice_date, layoff_date, location_text, *_ = cell_values
        else:
            employer, state_code, employee_count, notice_date, layoff_date, location_text, *_ = cell_values
        rows.append(
            {
                "employer": employer,
                "state_code": state_code,
                "employee_count": employee_count,
                "notice_date": parse_date_to_iso(notice_date),
                "effective_date": parse_date_to_iso(layoff_date),
                "location_text": location_text,
            }
        )
    return rows


def harvest_warntracker(fetcher, mapping_rows: list[dict]) -> list[dict]:
    rows = []
    latest_notice_date = ""
    for query in WARNTRACKER_QUERIES:
        query_url = WARNTRACKER_URL_TEMPLATE.format(query=urllib.parse.quote(query))
        html = fetcher.get(query_url, max_age_days=0.9).decode("utf-8", "replace")
        for parsed in parse_warntracker_rows(html):
            latest_notice_date = max(latest_notice_date, parsed["notice_date"])
            if not _is_recent(parsed["notice_date"]):
                continue
            state_full = STATE_ABBREV.get(_trim(parsed["state_code"]).upper(), _trim(parsed["state_code"]))
            matched = _matching_institution(parsed["employer"], state_full, mapping_rows)
            if matched is None:
                continue
            source_url = (
                f"{query_url}&employer={urllib.parse.quote(parsed['employer'])}"
                f"&state={urllib.parse.quote(parsed['state_code'])}"
                f"&notice_date={urllib.parse.quote(parsed['notice_date'])}"
            )
            rows.append(
                _build_warn_lead(
                    employer=parsed["employer"],
                    institution_row=matched,
                    source_url=source_url,
                    source_tag="warntracker_warn",
                    source_publication="WARNTracker",
                    state_full=state_full,
                    notice_date=parsed["notice_date"],
                    count_value=parsed["employee_count"],
                    event_text="Layoff",
                    effective_date=parsed["effective_date"],
                    location_text=parsed["location_text"],
                )
            )
    if latest_notice_date:
        _warn_if_stale("warntracker", latest_notice_date)
    return rows


def harvest_all(fetcher, mapping_rows: list[dict]) -> list[dict]:
    rows = []
    for source_name, runner in (
        ("ca", lambda: harvest_ca_warn(fetcher, mapping_rows)),
        ("wa", lambda: harvest_wa_warn(fetcher, mapping_rows)),
        ("ny", lambda: harvest_ny_warn(fetcher, mapping_rows)),
        ("warntracker", lambda: harvest_warntracker(fetcher, mapping_rows)),
    ):
        try:
            rows.extend(runner())
        except Exception as exc:
            print(f"::warning::cuts discovery WARN source {source_name} failed: {exc}")
    return rows
