"""Fixture-backed tests for WARN discovery harvesting."""

import io
import tempfile
from pathlib import Path
import sys

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "scripts"))

from cuts_discovery.harvest_warn import CA_WARN_URL, NY_WARN_URL, WA_WARN_URL, WARNTRACKER_URL_TEMPLATE, harvest_all  # noqa: E402
from import_supabase_institution_mapping import normalize_name  # noqa: E402


class FakeFetcher:
    def __init__(self, bodies: dict[str, bytes]):
        self.bodies = bodies

    def get(self, url: str, max_age_days: float = 6.0) -> bytes:  # noqa: ARG002
        return self.bodies[url]


def build_ca_workbook_bytes() -> bytes:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet("Index")
    workbook.create_sheet("WARN Report Summary")
    detailed = workbook.create_sheet("Detailed WARN Report ")
    detailed.append(["WARN REPORT", "", "", "", "", "", "", ""])
    detailed.append(["County/Parish", "Notice Date", "Processed Date", "Effective Date", "Company", "Layoff/Closure", "No. Of Employees", "Address"])
    detailed.append(["Orange County", "2026-07-08", "2026-07-09", "2026-08-15", "California Example University", "Layoff Permanent", 42, "1 Campus Way Irvine CA 92612"])
    detailed.append(["Orange County", "2025-01-01", "2025-01-02", "2025-02-01", "Old Example University", "Layoff Permanent", 9, "2 Old Way Irvine CA 92612"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def main():
    mapping_rows = [
        {
            "norm_name": normalize_name("California Example University"),
            "unitid": "100",
            "state": "California",
            "institution_name": "California Example University",
            "match_method": "test",
        },
        {
            "norm_name": normalize_name("Washington Example College"),
            "unitid": "200",
            "state": "Washington",
            "institution_name": "Washington Example College",
            "match_method": "test",
        },
        {
            "norm_name": normalize_name("The College of Saint Rose"),
            "unitid": "300",
            "state": "New York",
            "institution_name": "The College of Saint Rose",
            "match_method": "test",
        },
    ]

    wa_html = """
    <html><body>
      <table id="ucPSW_gvMain">
        <tr><th>Employer</th></tr>
        <tr>
          <td>Washington Example College</td>
          <td>Seattle</td>
          <td>8/15/2026</td>
          <td>17</td>
          <td>Layoff</td>
          <td>Permanent</td>
          <td>7/10/2026</td>
          <td><a href='/esd/file/WARN/Public/DownloadFile.aspx?file=wa-example.pdf&download=1'>Download</a></td>
        </tr>
      </table>
    </body></html>
    """
    ny_html = """
    <html><body>
      <table>
        <tr>
          <td><a href="/warn-college-saint-rose-region-capital-date-notice-2152025-original-notice-date-2152024-date-posted">The College of Saint Rose</a></td>
          <td>Capital</td>
          <td>7/11/2026</td>
          <td>7/10/2026</td>
        </tr>
        <tr>
          <td><a href="/warn-generic-company">Generic Company</a></td>
          <td>Capital</td>
          <td>7/11/2026</td>
          <td>7/10/2026</td>
        </tr>
      </table>
    </body></html>
    """
    ny_notice_text = """
    NEW YORK STATE DEPARTMENT OF LABOR
    WARN UNIT
    Reason For Closure: Economic
    Company:
    The College of Saint Rose
    FEIN NUM: 14-1338371
    Total Number of Affected Workers: 646
    Closure End Date: July 31, 2026
    Number of Affected Employees at Site: 646
    Date of Notice: July 10, 2026
    """
    warntracker_html = """
    <html><body>
      <table>
        <tbody>
          <tr>
            <td>1</td>
            <td>Aggregator University</td>
            <td>CA</td>
            <td>23</td>
            <td>7/09/2026</td>
            <td>8/01/2026</td>
            <td>Los Angeles</td>
            <td>2026</td>
          </tr>
        </tbody>
      </table>
    </body></html>
    """

    fetcher = FakeFetcher(
        bodies={
            CA_WARN_URL: build_ca_workbook_bytes(),
            WA_WARN_URL: wa_html.encode("utf-8"),
            NY_WARN_URL: ny_html.encode("utf-8"),
            "https://dol.ny.gov/warn-college-saint-rose-region-capital-date-notice-2152025-original-notice-date-2152024-date-posted": ny_notice_text.encode("utf-8"),
            WARNTRACKER_URL_TEMPLATE.format(query="College"): b"<html><body><table><tbody><tr><td colSpan='8'>No records to display</td></tr></tbody></table></body></html>",
            WARNTRACKER_URL_TEMPLATE.format(query="University"): warntracker_html.encode("utf-8"),
            WARNTRACKER_URL_TEMPLATE.format(query="Institute"): b"<html><body><table><tbody><tr><td colSpan='8'>No records to display</td></tr></tbody></table></body></html>",
            WARNTRACKER_URL_TEMPLATE.format(query="Seminary"): b"<html><body><table><tbody><tr><td colSpan='8'>No records to display</td></tr></tbody></table></body></html>",
        }
    )

    mapping_rows.append(
        {
            "norm_name": normalize_name("Aggregator University"),
            "unitid": "400",
            "state": "California",
            "institution_name": "Aggregator University",
            "match_method": "test",
        }
    )

    rows = harvest_all(fetcher, mapping_rows)
    assert len(rows) == 4, rows
    by_source = {row["query_or_feed"]: row for row in rows}
    assert set(by_source) == {"ca_warn", "wa_warn", "ny_warn", "warntracker_warn"}, by_source

    assert by_source["ca_warn"]["headline"] == "California Example University WARN layoff notice affecting 42 employees", by_source["ca_warn"]
    assert "Affects 42 employees." in by_source["ca_warn"]["snippet"], by_source["ca_warn"]
    assert by_source["wa_warn"]["published_date"] == "2026-07-10", by_source["wa_warn"]
    assert "Washington WARN" == by_source["wa_warn"]["publisher"], by_source["wa_warn"]
    assert by_source["ny_warn"]["published_date"] == "2026-07-10", by_source["ny_warn"]
    assert "Affects 646 employees." in by_source["ny_warn"]["snippet"], by_source["ny_warn"]
    assert by_source["warntracker_warn"]["publisher"] == "WARNTracker", by_source["warntracker_warn"]
    assert by_source["warntracker_warn"]["published_date"] == "2026-07-09", by_source["warntracker_warn"]
    assert all(row["tier"] == "warn" for row in rows), rows

    print("cuts discovery WARN harvest tests: all passed.")


if __name__ == "__main__":
    main()
